"""
HUBMAIS MANAGER — Painel de Métricas
=======================================
Lê a Base Histórica (gerada pelo atualizar_base.py) e calcula as métricas
derivadas: individual (colaborador) e supervisor (visão do time).

Gera dashboard.xlsx com 2 abas: "Painel Individual" e "Painel Supervisor".

⚠️ PONTOS QUE DEPENDEM DE VOCÊ CONFIRMAR (ajustar no bloco CONFIG abaixo):
  - META_SEMANAL_RECEBIDO: hoje é um valor-suposição (5 dias úteis x R$ 2.200,
    que é o piso da faixa "Meta cumprida" no Resultado Parcial). Se a meta
    semanal real for outra, troca aqui.
  - LIMIAR_TMA_FORA_PADRAO: % de desvio da média do turno pra marcar alguém
    com TMA fora do padrão (hoje 30% pra mais ou pra menos).
"""

from pathlib import Path
from calendar import monthrange
from datetime import timedelta

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# CONFIG (ajustável)
# ==========================================
PASTA_PROJETO = Path(__file__).resolve().parent.parent
ARQUIVO_BASE = PASTA_PROJETO / "dados" / "base_historica.xlsx"
ARQUIVO_DASHBOARD = PASTA_PROJETO / "dados" / "dashboard.xlsx"

META_DIARIA_RECEBIDO = 1923
DIAS_UTEIS_SEMANA = 6              # seg a sáb
META_SEMANAL_RECEBIDO = META_DIARIA_RECEBIDO * DIAS_UTEIS_SEMANA  # R$ 11.538
LIMIAR_TMA_FORA_PADRAO = 0.30      # 30% de desvio da média do turno

# Faixas de premiação — % da meta mensal batido (ordem decrescente) → bônus em R$.
# Ex.: meta mensal R$50.000 → 100% = R$50.000 (300), 75% = R$37.500 (200), 60% = R$30.000 (100).
# As faixas são em % da meta pra continuarem corretas mesmo quando a meta mensal
# mudar de valor (ela varia com o nº de dias úteis do mês).
FAIXAS_PREMIACAO = [
    (1.00, 300),
    (0.75, 200),
    (0.60, 100),
]
LIMIAR_CONVERSAO_PREMIACAO = 0.65  # % recebido/negociado mínimo no mês pra ter direito à premiação


def dias_uteis_no_mes(ano: int, mes: int) -> int:
    """Conta quantos dias seg-sáb (dia útil da empresa) tem no mês inteiro."""
    total_dias = monthrange(ano, mes)[1]
    return sum(1 for d in range(1, total_dias + 1) if pd.Timestamp(ano, mes, d).weekday() != 6)


def dias_uteis_restantes(data_referencia: pd.Timestamp) -> int:
    """Quantos dias seg-sáb faltam no mês, contando a partir do dia SEGUINTE à referência."""
    ultimo_dia = monthrange(data_referencia.year, data_referencia.month)[1]
    restantes = 0
    for d in range(data_referencia.day + 1, ultimo_dia + 1):
        if pd.Timestamp(data_referencia.year, data_referencia.month, d).weekday() != 6:
            restantes += 1
    return restantes


def calcular_premiacao(recebido_mes: float, meta_mensal: float, negociado_mes: float) -> dict:
    """Calcula a faixa de premiação atual, a próxima, e o ritmo necessário pra alcançá-la."""
    pct_conversao_mes = (recebido_mes / negociado_mes) if negociado_mes else None
    elegivel_conversao = pct_conversao_mes is not None and pct_conversao_mes >= LIMIAR_CONVERSAO_PREMIACAO

    pct_meta = recebido_mes / meta_mensal if meta_mensal else 0
    faixa_atual = None
    for pct_faixa, valor in FAIXAS_PREMIACAO:  # já em ordem decrescente
        if pct_meta >= pct_faixa:
            faixa_atual = (pct_faixa, valor)
            break

    faixas_nao_atingidas = [f for f in reversed(FAIXAS_PREMIACAO) if faixa_atual is None or f[0] > faixa_atual[0]]
    proxima_faixa = faixas_nao_atingidas[0] if faixas_nao_atingidas else None

    return {
        "elegivel_conversao": elegivel_conversao,
        "pct_conversao_mes": pct_conversao_mes,
        "valor_bonus_atual": faixa_atual[1] if faixa_atual else 0,
        "pct_faixa_atual": faixa_atual[0] if faixa_atual else None,
        "proxima_faixa_valor": proxima_faixa[1] if proxima_faixa else None,
        "falta_proxima_faixa": max((proxima_faixa[0] * meta_mensal) - recebido_mes, 0) if proxima_faixa else 0,
    }


# ==========================================
# HELPERS
# ==========================================
def tma_para_segundos(valor) -> float:
    """Converte 'HH:MM:SS' (string salva pelo atualizar_base.py) em segundos."""
    if pd.isna(valor) or valor in (None, "", "-"):
        return None
    h, m, s = str(valor).split(":")
    return int(h) * 3600 + int(m) * 60 + int(s)


def segundos_para_tma(segundos) -> str:
    if pd.isna(segundos) or segundos is None:
        return "-"
    total = int(round(segundos))
    h, resto = divmod(total, 3600)
    m, s = divmod(resto, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


def carregar_base() -> pd.DataFrame:
    df = pd.read_excel(ARQUIVO_BASE, sheet_name="Base Histórica")
    df["Data"] = pd.to_datetime(df["Data"])
    df["TMA_segundos"] = df["TMA"].apply(tma_para_segundos)
    iso = df["Data"].dt.isocalendar()
    df["Semana"] = iso.week
    df["AnoSemana"] = iso.year  # evita ambiguidade em virada de ano
    return df


def dias_fechados_ordenados(df: pd.DataFrame):
    return sorted(df.dropna(subset=["Recebido"])["Data"].unique())


# ==========================================
# MÉTRICAS INDIVIDUAIS (calculadas "como se hoje fosse" data_referencia)
# ==========================================
def calcular_individual(df: pd.DataFrame, data_referencia=None) -> pd.DataFrame:
    dias_fechados = dias_fechados_ordenados(df)
    if not dias_fechados:
        raise ValueError("Nenhum dia fechado (com Recebido) encontrado na base.")

    data_hoje = pd.Timestamp(data_referencia) if data_referencia is not None else dias_fechados[-1]
    anteriores = [d for d in dias_fechados if d < data_hoje]
    data_ontem = anteriores[-1] if anteriores else None

    linha_ref = df.loc[df["Data"] == data_hoje]
    semana_atual = linha_ref["Semana"].iloc[0]
    anosemana_atual = linha_ref["AnoSemana"].iloc[0]
    semana_anterior_data = data_hoje - timedelta(days=7)
    ano_mes_atual = (data_hoje.year, data_hoje.month)
    meta_mensal_total = META_DIARIA_RECEBIDO * dias_uteis_no_mes(*ano_mes_atual)

    linhas = []
    for colaborador, grupo in df.groupby("Colaborador"):
        hoje = grupo[grupo["Data"] == data_hoje]
        ontem = grupo[grupo["Data"] == data_ontem] if data_ontem is not None else grupo.iloc[0:0]
        semana = grupo[(grupo["Semana"] == semana_atual) & (grupo["AnoSemana"] == anosemana_atual) & (grupo["Data"] <= data_hoje)]
        semana_ant = grupo[
            (grupo["Data"] > semana_anterior_data - timedelta(days=7))
            & (grupo["Data"] <= semana_anterior_data)
        ]
        mes = grupo[
            (grupo["Data"].dt.year == ano_mes_atual[0])
            & (grupo["Data"].dt.month == ano_mes_atual[1])
            & (grupo["Data"] <= data_hoje)
        ]

        recebido_hoje = hoje["Recebido"].sum() if not hoje.empty else 0.0
        negociado_hoje = hoje["Negociado"].sum() if not hoje.empty else 0.0
        recebido_ontem = ontem["Recebido"].sum() if not ontem.empty else None
        recebido_semana = semana["Recebido"].sum()
        recebido_semana_ant = semana_ant["Recebido"].sum() if not semana_ant.empty else None
        recebido_mes = mes["Recebido"].sum()
        negociado_mes = mes["Negociado"].sum()
        tma_hoje = hoje["TMA_segundos"].iloc[0] if not hoje.empty else None

        pct_recebimento = (recebido_hoje / negociado_hoje) if negociado_hoje else None
        progresso_meta_semana = recebido_semana / META_SEMANAL_RECEBIDO
        progresso_meta_mes = recebido_mes / meta_mensal_total

        evolucao = (recebido_hoje - recebido_ontem) / recebido_ontem if recebido_ontem else None
        evolucao_semana = (
            (recebido_semana - recebido_semana_ant) / recebido_semana_ant
            if recebido_semana_ant else None
        )

        premiacao = calcular_premiacao(recebido_mes, meta_mensal_total, negociado_mes)

        linhas.append({
            "Colaborador": colaborador,
            "Data": data_hoje,
            "TMA do dia": segundos_para_tma(tma_hoje),
            "Recebido hoje": recebido_hoje,
            "% Recebimento (recebido/negociado)": pct_recebimento,
            "Recebido na semana": recebido_semana,
            "Meta semanal": META_SEMANAL_RECEBIDO,
            "Progresso da meta semanal": progresso_meta_semana,
            "Recebido na semana anterior": recebido_semana_ant,
            "Evolução vs. semana anterior": evolucao_semana,
            "Recebido no mês": recebido_mes,
            "Negociado no mês": negociado_mes,
            "Meta mensal": meta_mensal_total,
            "Progresso da meta mensal": progresso_meta_mes,
            "Evolução vs. dia anterior": evolucao,
            "Dias úteis restantes no mês": dias_uteis_restantes(data_hoje),
            **{f"Premiação: {k}": v for k, v in premiacao.items()},
        })

    resultado = pd.DataFrame(linhas)
    resultado["Ranking do turno (recebido hoje)"] = (
        resultado["Recebido hoje"].rank(ascending=False, method="min").astype(int)
    )
    resultado = resultado.sort_values("Ranking do turno (recebido hoje)").reset_index(drop=True)
    return resultado


# ==========================================
# MÉTRICAS DE SUPERVISOR
# ==========================================
def calcular_supervisor(individual: pd.DataFrame, df: pd.DataFrame) -> pd.DataFrame:
    v = individual.copy()

    v["Ranking recebimento"] = v["Recebido hoje"].rank(ascending=False, method="min").astype("Int64")
    v["Ranking conversão"] = v["% Recebimento (recebido/negociado)"].rank(ascending=False, method="min").astype("Int64")

    media_tma = v["TMA do dia"].apply(tma_para_segundos).mean()
    v["TMA fora do padrão?"] = v["TMA do dia"].apply(tma_para_segundos).apply(
        lambda s: "⚠️ Sim" if pd.notna(s) and media_tma and abs(s - media_tma) / media_tma > LIMIAR_TMA_FORA_PADRAO
        else "-"
    )

    v["Melhorando / Caindo"] = v["Evolução vs. dia anterior"].apply(
        lambda e: "📈 Melhorando" if pd.notna(e) and e > 0
        else ("📉 Caindo" if pd.notna(e) and e < 0 else "-")
    )

    # Negocia muito x recebe pouco: acima da mediana em negociado, abaixo da mediana em conversão
    mediana_negociado = individual["Recebido na semana"].median()  # proxy de volume de atividade
    mediana_conv = individual["% Recebimento (recebido/negociado)"].median()
    negociado_hoje_map = df.groupby("Colaborador")["Negociado"].sum()
    v["Negocia muito x recebe pouco?"] = v.apply(
        lambda r: "⚠️ Sim" if (negociado_hoje_map.get(r["Colaborador"], 0) > negociado_hoje_map.median())
        and pd.notna(r["% Recebimento (recebido/negociado)"])
        and r["% Recebimento (recebido/negociado)"] < mediana_conv
        else "-",
        axis=1,
    )

    negociacoes_map = df.groupby("Colaborador")["Negociações"].sum()
    v["Recebe bem x baixo volume?"] = v.apply(
        lambda r: "⚠️ Sim" if r["Recebido hoje"] > v["Recebido hoje"].median()
        and negociacoes_map.get(r["Colaborador"], 0) < negociacoes_map.median()
        else "-",
        axis=1,
    )

    colunas = ["Colaborador", "Recebido hoje", "Ranking recebimento",
               "% Recebimento (recebido/negociado)", "Ranking conversão",
               "TMA do dia", "TMA fora do padrão?", "Melhorando / Caindo",
               "Negocia muito x recebe pouco?", "Recebe bem x baixo volume?"]
    return v[colunas].sort_values("Ranking recebimento").reset_index(drop=True)


# ==========================================
# SUGESTÕES AUTOMÁTICAS (baseadas em padrão dos números agregados)
# ==========================================
def gerar_sugestao(linha: dict, sup: dict) -> str:
    """
    Sugestão de ação pro colaborador, priorizando o sinal mais urgente.
    IMPORTANTE: isso é heurística sobre os números agregados (recebido, negociado,
    TMA, conversão) — não temos dado de cliente individual nem de histórico de
    promessas de pagamento (isso está no CRM/discador, fora do escopo desta base).
    A sugestão aponta o padrão, não nomeia cliente nem promessa específica.
    """
    evo_sem = linha["Evolução vs. semana anterior"]
    evo_dia = linha["Evolução vs. dia anterior"]
    conv_mes = linha["Premiação: pct_conversao_mes"]
    elegivel = linha["Premiação: elegivel_conversao"]

    if conv_mes is not None and not elegivel:
        return (f"Conversão do mês em {conv_mes*100:.1f}%, abaixo dos "
                f"{LIMIAR_CONVERSAO_PREMIACAO*100:.0f}% exigidos pra premiação — "
                "priorizar cobrar quem já negociou e ainda não pagou antes de fechar acordo novo.")

    if evo_sem is not None and evo_sem <= -0.20:
        return (f"Recebido caiu {abs(evo_sem)*100:.0f}% em relação à semana passada — "
                "vale retomar contato com quem prometeu pagar essa semana e não pagou.")

    if sup["Negocia muito x recebe pouco?"] != "-":
        return ("Fecha bastante negociação mas pouco disso vira pagamento — "
                "reforçar confirmação da promessa (dia, forma de pagamento) antes de considerar fechado.")

    if sup["Recebe bem x baixo volume?"] != "-":
        return "Boa conversão no que negocia, mas baixo volume — vale aumentar o número de contatos ativos no dia."

    if sup["TMA fora do padrão?"] != "-":
        return "TMA fora do padrão do turno — revisar abordagem da ligação (tempo muito longo ou curto demais)."

    if evo_dia is not None and evo_dia <= -0.30:
        return f"Recebido caiu {abs(evo_dia)*100:.0f}% vs. o dia anterior — vale atenção nos próximos contatos."

    if linha["Premiação: valor_bonus_atual"] >= 300:
        return "Na faixa máxima de premiação e com conversão em dia — manter a rotina atual."

    return "Sem sinal de alerta relevante hoje — manter a rotina de contatos."


# ==========================================
# ESCRITA DO EXCEL
# ==========================================
def formatar_aba(ws, df: pd.DataFrame, colunas_pct=(), colunas_moeda=(), larguras=None):
    fonte_padrao = Font(name="Arial", size=10)
    fonte_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    fill_header = PatternFill("solid", fgColor="1F4E78")
    borda_fina = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

    ws.append(list(df.columns))
    for cel in ws[1]:
        cel.font = fonte_header
        cel.fill = fill_header
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32

    for _, linha in df.iterrows():
        ws.append(list(linha))

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for idx, cel in enumerate(row):
            nome_col = df.columns[idx]
            cel.font = fonte_padrao
            cel.border = borda_fina
            if nome_col in colunas_pct:
                cel.number_format = "0.0%"
            elif nome_col in colunas_moeda:
                cel.number_format = 'R$ #,##0.00;(R$ #,##0.00);"-"'

    larguras = larguras or [18] * len(df.columns)
    for i, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = largura

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def salvar_dashboard(individual: pd.DataFrame, supervisor: pd.DataFrame, caminho: Path):
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Painel Individual"
    formatar_aba(
        ws1, individual,
        colunas_pct=("% Recebimento (recebido/negociado)", "Progresso da meta semanal",
                     "Progresso da meta mensal", "Evolução vs. dia anterior", "Evolução vs. semana anterior",
                     "Premiação: pct_conversao_mes", "Premiação: pct_faixa_atual"),
        colunas_moeda=("Recebido hoje", "Recebido na semana", "Meta semanal",
                        "Recebido na semana anterior", "Recebido no mês", "Negociado no mês", "Meta mensal",
                        "Premiação: valor_bonus_atual", "Premiação: proxima_faixa_valor", "Premiação: falta_proxima_faixa"),
    )

    ws2 = wb.create_sheet("Painel Supervisor")
    formatar_aba(
        ws2, supervisor,
        colunas_pct=("% Recebimento (recebido/negociado)",),
        colunas_moeda=("Recebido hoje",),
        larguras=[30, 14, 16, 24, 16, 12, 16, 18, 20, 20],
    )

    wb.save(caminho)


# ==========================================
# EXECUÇÃO
# ==========================================
if __name__ == "__main__":
    print("=" * 60)
    print("📊 HUBMAIS MANAGER — PAINEL DE MÉTRICAS")
    print("=" * 60)

    df = carregar_base()
    individual = calcular_individual(df)
    supervisor = calcular_supervisor(individual, df)

    print(f"\n✅ Métricas calculadas para {len(individual)} colaboradores")
    print(f"   Data de referência: {individual['Data'].iloc[0].date()}")

    try:
        salvar_dashboard(individual, supervisor, ARQUIVO_DASHBOARD)
    except PermissionError:
        print(f"\n❌ Não consegui salvar: '{ARQUIVO_DASHBOARD.name}' parece estar aberto no Excel.")
        raise SystemExit(1)

    print(f"\n💾 Dashboard salvo em: {ARQUIVO_DASHBOARD}")
