"""
HUBMAIS MANAGER — Atualização da Base Central
================================================
Lê o relatório de TMA (chamadas) e o Resultado Parcial (negociação/recebimento),
cruza os dois pelos 7 colaboradores do turno Vespertino / Sup. Reginaldo,
e grava/atualiza a Base Histórica única.

Estrutura da Base Central:
Data | Colaborador | Supervisor | Turno | TMA | Negociado | Recebido | Negociações | Clientes Pagantes
"""

import re
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# CONFIGURAÇÕES
# ==========================================
PASTA_PROJETO = Path(__file__).resolve().parent.parent
PASTA_TMA = PASTA_PROJETO / "dados" / "tma"
PASTA_RESULTADOS = PASTA_PROJETO / "dados" / "resultados"
ARQUIVO_BASE = PASTA_PROJETO / "dados" / "base_historica.xlsx"

SUPERVISOR = "Reginaldo Sousa Barbosa"
TURNO = "Vespertino"

# Mapeamento validado nos dois relatórios reais (19/08 e 18/08/2026).
# tma_id   = "ID Sistema" na aba Diário/Consolidado do relatório de TMA
# login    = "Login Abridor Negociação" / "NOME_LOGIN" no Resultado Parcial
COLABORADORES = {
    "ALEXANDRE LOPES DOS SANTOS":     {"tma_id": "alexandresantos",  "login": "ALSANTOS"},
    "IVONEIDE DUARTE FERREIRA":       {"tma_id": "ivoneideferreira", "login": "IDFERREIRA"},
    "JOAO PEDRO DE SOUZA SANTOS":     {"tma_id": "joaosantos",       "login": "JPDSOUZA"},
    "KALISSA VALERIO GONCALVES":      {"tma_id": "kalissagoncalves", "login": "KGONCALVES"},
    "MARIA DO CARMO SOUZA SANTOS":    {"tma_id": "mariasantos",      "login": "MCSSANTOS"},
    "NAYARA BISPO DO ESPIRITO SANTO": {"tma_id": "nayarapereira",    "login": "NSPEREIRA"},
    "YARA DO CARMO DA CONCEICAO":     {"tma_id": "yaraconceicao",    "login": "YCONCEIÇÃO"},
}

NOMES_PROPRIOS = {nome: nome.title() for nome in COLABORADORES}
LOGIN_PARA_NOME = {v["login"].upper(): k for k, v in COLABORADORES.items()}


# ==========================================
# HELPERS
# ==========================================
def normalizar(texto: str) -> str:
    """Maiúsculas, sem acento, sem espaço extra — para comparar nomes/logins com segurança."""
    if texto is None:
        return ""
    texto = str(texto).strip().upper()
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", texto)


def buscar_ultimo_arquivo(pasta: Path) -> Path:
    arquivos = list(pasta.glob("*.xlsx"))
    if not arquivos:
        raise FileNotFoundError(f"Nenhum arquivo Excel encontrado em: {pasta}")
    return max(arquivos, key=lambda a: a.stat().st_mtime)


# ==========================================
# TMA — aba "Diário" (uma linha por colaborador por dia)
# ==========================================
def processar_tma(caminho: Path) -> pd.DataFrame:
    # A aba tem um título mesclado na linha 1 ("Pausas Detalhadas");
    # o cabeçalho de verdade está na linha 2.
    df = pd.read_excel(caminho, sheet_name="Diário", header=1)
    df["nome_norm"] = df["Nome Completo"].apply(normalizar)

    nomes_validos = set(COLABORADORES.keys())
    df = df[df["nome_norm"].isin(nomes_validos)].copy()

    df["Colaborador"] = df["nome_norm"].map(NOMES_PROPRIOS)
    df["Data"] = pd.to_datetime(df["Data"]).dt.normalize()
    df["TMA_segundos"] = df["TMA"].apply(
        lambda x: x.total_seconds() if pd.notnull(x) else None
    )

    return df[["Data", "Colaborador", "TMA_segundos"]]


# ==========================================
# RESULTADO PARCIAL — abas em formato matriz (1 coluna por dia do mês)
# ==========================================
def melt_aba_resultado(caminho: Path, aba: str, nome_valor: str, ano: int, mes: int) -> pd.DataFrame:
    df = pd.read_excel(caminho, sheet_name=aba)
    col_id = df.columns[0]

    # mantém só colunas de dia (inteiros); descarta "Total" e "Unnamed: N"
    col_dias = [c for c in df.columns[1:] if isinstance(c, (int, float)) and not pd.isna(c)]

    df = df[[col_id] + col_dias].rename(columns={col_id: "login"})
    df["login_norm"] = df["login"].apply(normalizar)

    logins_validos = {normalizar(v["login"]) for v in COLABORADORES.values()}
    df = df[df["login_norm"].isin(logins_validos)]

    derretido = df.melt(
        id_vars=["login_norm"], value_vars=col_dias, var_name="dia", value_name=nome_valor
    )
    derretido["Data"] = derretido["dia"].apply(lambda d: datetime(ano, mes, int(d)))
    return derretido[["login_norm", "Data", nome_valor]]


def processar_resultado(caminho: Path, ano: int, mes: int) -> pd.DataFrame:
    negociado = melt_aba_resultado(caminho, "NEGOCIAÇÃO_Valor por dia", "Negociado", ano, mes)
    recebido = melt_aba_resultado(caminho, "RECEBIMENTO_Valor", "Recebido", ano, mes)
    negociacoes = melt_aba_resultado(caminho, "NEGOCIAÇÃO_Qtde Faturas", "Negociacoes", ano, mes)
    clientes_pag = melt_aba_resultado(caminho, "RECEBIMENTO_Qtde Clientes", "ClientesPagantes", ano, mes)

    base = negociado.merge(recebido, on=["login_norm", "Data"], how="outer")
    base = base.merge(negociacoes, on=["login_norm", "Data"], how="outer")
    base = base.merge(clientes_pag, on=["login_norm", "Data"], how="outer")

    login_norm_para_nome = {normalizar(v["login"]): k for k, v in COLABORADORES.items()}
    base["Colaborador"] = base["login_norm"].map(login_norm_para_nome).map(NOMES_PROPRIOS)

    return base[["Data", "Colaborador", "Negociado", "Recebido", "Negociacoes", "ClientesPagantes"]]


# ==========================================
# MONTAGEM DA BASE CENTRAL
# ==========================================
def montar_base_central(df_tma: pd.DataFrame, df_resultado: pd.DataFrame) -> pd.DataFrame:
    # Último dia realmente coberto pelo Resultado Parcial. Dias além disso ainda não
    # fecharam no sistema de negociação/recebimento — ficam em branco (pendente),
    # não em zero, para não parecer "colaborador não negociou nada".
    data_limite_resultado = df_resultado["Data"].max()

    base = df_tma.merge(df_resultado, on=["Data", "Colaborador"], how="outer")

    base["Supervisor"] = SUPERVISOR
    base["Turno"] = TURNO

    base = base.rename(columns={
        "TMA_segundos": "TMA (segundos)",
        "Negociacoes": "Negociações",
        "ClientesPagantes": "Clientes Pagantes",
    })

    # Dentro do período já fechado pelo Resultado Parcial, dias sem coluna
    # correspondente (ex.: domingos, onde a aba de Negociação às vezes nem lista
    # o dia) significam ausência de atividade = 0, não "sem dado". Além desse
    # período, mantém em branco (pendente de fechamento).
    dentro_do_periodo = base["Data"] <= data_limite_resultado
    for col in ["Negociado", "Recebido", "Negociações", "Clientes Pagantes"]:
        base.loc[dentro_do_periodo, col] = base.loc[dentro_do_periodo, col].fillna(0)

    colunas = ["Data", "Colaborador", "Supervisor", "Turno", "TMA (segundos)",
               "Negociado", "Recebido", "Negociações", "Clientes Pagantes"]
    base = base[colunas].sort_values(["Data", "Colaborador"]).reset_index(drop=True)
    return base


def carregar_base_existente() -> pd.DataFrame:
    if not ARQUIVO_BASE.exists() or ARQUIVO_BASE.stat().st_size == 0:
        return pd.DataFrame()
    try:
        return pd.read_excel(ARQUIVO_BASE, sheet_name="Base Histórica")
    except Exception as erro:
        print(f"⚠️  Não consegui ler a base existente ({erro}).")
        print("    Pode estar aberta no Excel, corrompida ou vazia — tratando como inexistente.")
        return pd.DataFrame()


def atualizar_base_historica(base_nova: pd.DataFrame) -> pd.DataFrame:
    """Anexa a base nova à existente, sem duplicar linhas de Data+Colaborador já gravadas
    (a linha nova sempre substitui a antiga, pois pode vir com dado mais atualizado)."""
    base_antiga = carregar_base_existente()
    if base_antiga.empty:
        return base_nova

    chave = ["Data", "Colaborador"]
    base_antiga["_chave"] = list(zip(base_antiga["Data"], base_antiga["Colaborador"]))
    base_nova["_chave"] = list(zip(base_nova["Data"], base_nova["Colaborador"]))

    base_antiga = base_antiga[~base_antiga["_chave"].isin(base_nova["_chave"])]

    final = pd.concat([base_antiga, base_nova], ignore_index=True)
    final = final.drop(columns="_chave").sort_values(chave).reset_index(drop=True)
    return final


# ==========================================
# ESCRITA DO EXCEL (formatado)
# ==========================================
def formatar_tma(segundos):
    if pd.isna(segundos):
        return None
    total = int(round(segundos))
    h, resto = divmod(total, 3600)
    m, s = divmod(resto, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


def salvar_excel(base: pd.DataFrame, caminho: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Base Histórica"

    fonte_padrao = Font(name="Arial", size=10)
    fonte_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    fill_header = PatternFill("solid", fgColor="1F4E78")
    borda_fina = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

    colunas = ["Data", "Colaborador", "Supervisor", "Turno", "TMA",
               "Negociado", "Recebido", "Negociações", "Clientes Pagantes"]
    ws.append(colunas)
    for cel in ws[1]:
        cel.font = fonte_header
        cel.fill = fill_header
        cel.alignment = Alignment(horizontal="center", vertical="center")

    for _, linha in base.iterrows():
        ws.append([
            linha["Data"],
            linha["Colaborador"],
            linha["Supervisor"],
            linha["Turno"],
            formatar_tma(linha["TMA (segundos)"]),
            linha["Negociado"] if pd.notna(linha["Negociado"]) else None,
            linha["Recebido"] if pd.notna(linha["Recebido"]) else None,
            linha["Negociações"] if pd.notna(linha["Negociações"]) else None,
            linha["Clientes Pagantes"] if pd.notna(linha["Clientes Pagantes"]) else None,
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[0].number_format = "DD/MM/YYYY"
        row[5].number_format = 'R$ #,##0.00;(R$ #,##0.00);"-"'
        row[6].number_format = 'R$ #,##0.00;(R$ #,##0.00);"-"'
        for cel in row:
            cel.font = fonte_padrao
            cel.border = borda_fina

    larguras = [12, 32, 26, 12, 10, 14, 14, 13, 16]
    for i, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = largura

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(caminho)


# ==========================================
# EXECUÇÃO
# ==========================================
if __name__ == "__main__":
    print("=" * 60)
    print("🚀 HUBMAIS MANAGER — ATUALIZAÇÃO DA BASE CENTRAL")
    print("=" * 60)

    arquivo_tma = buscar_ultimo_arquivo(PASTA_TMA)
    arquivo_resultado = buscar_ultimo_arquivo(PASTA_RESULTADOS)
    print(f"📂 TMA:       {arquivo_tma.name}")
    print(f"📂 Resultado: {arquivo_resultado.name}")

    df_tma = processar_tma(arquivo_tma)
    df_resultado = processar_resultado(arquivo_resultado, ano=2026, mes=8)

    print(f"\n✅ TMA processado: {len(df_tma)} linhas (colaborador x dia)")
    print(f"✅ Resultado processado: {len(df_resultado)} linhas (colaborador x dia)")

    base_nova = montar_base_central(df_tma, df_resultado)
    base_final = atualizar_base_historica(base_nova)

    print(f"\n✅ Base Central final: {len(base_final)} linhas")
    print(f"   Período: {base_final['Data'].min().date()} até {base_final['Data'].max().date()}")

    try:
        salvar_excel(base_final, ARQUIVO_BASE)
    except PermissionError:
        print(f"\n❌ Não consegui salvar: '{ARQUIVO_BASE.name}' parece estar aberto no Excel.")
        print("   Feche o arquivo e rode o script de novo.")
        raise SystemExit(1)

    print(f"\n💾 Base salva em: {ARQUIVO_BASE}")